"""
Tool Functions for My Hiking Chatbot
=====================================
Functions called by Gemini via function calling:
booking, SAR dashboard, Excel export, CRUD mountain/trail.
"""

import os
import json
import datetime
import requests

from config import LARAVEL_API_URL, EXPORT_DIR, CHATBOT_SECRET
from database import (
    fetch_trails_data,
    fetch_panic_data,
    fetch_mountains_data,
    fetch_orders_data,
    fetch_transactions_data,
    fetch_users_data,
    fetch_trails_by_guard,
    fetch_orders_by_trail_ids,
    fetch_panics_by_trail_ids,
    fetch_transactions_by_trail_ids,
)


# ============================================
# HELPER / NORMALIZER FUNCTIONS
# ============================================

def _normalize_member_ids(raw_member_ids, user_id=None):
    """Normalisasi data anggota agar selalu list[int] unik dan valid."""
    if raw_member_ids is None:
        return []

    if isinstance(raw_member_ids, (int, str)):
        raw_member_ids = [raw_member_ids]
    elif not isinstance(raw_member_ids, list):
        return []

    normalized = []
    for uid in raw_member_ids:
        try:
            parsed_uid = int(uid)
            if parsed_uid <= 0:
                continue
            if user_id is not None and parsed_uid == int(user_id):
                continue
            normalized.append(parsed_uid)
        except Exception:
            continue

    return list(dict.fromkeys(normalized))


def _normalize_positive_int(raw_value):
    """Konversi nilai numerik ke int positif (mendukung 1, "1", 1.0)."""
    if raw_value is None or isinstance(raw_value, bool):
        return None

    try:
        if isinstance(raw_value, int):
            parsed = raw_value
        elif isinstance(raw_value, float):
            if not raw_value.is_integer():
                return None
            parsed = int(raw_value)
        elif isinstance(raw_value, str):
            text = raw_value.strip()
            if not text:
                return None
            if '.' in text:
                as_float = float(text)
                if not as_float.is_integer():
                    return None
                parsed = int(as_float)
            else:
                parsed = int(text)
        else:
            parsed = int(raw_value)
    except Exception:
        return None

    return parsed if parsed > 0 else None


def _laravel_json_headers(auth_token=None, chatbot_crud=False):
    """Bangun header JSON + Bearer token (jika tersedia).

    Args:
        auth_token: Sanctum bearer token untuk endpoint yang memerlukan auth user.
        chatbot_crud: Jika True, tambahkan header X-Chatbot-Secret untuk
                      endpoint CRUD chatbot admin di Laravel.
    """
    headers = {'Content-Type': 'application/json'}
    token = (auth_token or '').strip()
    if token.lower().startswith('bearer '):
        token = token[7:].strip()
    if token:
        headers['Authorization'] = f'Bearer {token}'
    if chatbot_crud and CHATBOT_SECRET:
        headers['X-Chatbot-Secret'] = CHATBOT_SECRET
    return headers


def _safe_json_from_response(response):
    """Ambil JSON response secara aman tanpa melempar exception."""
    try:
        return response.json()
    except Exception:
        return {}


def _humanize_profile_field(field_name):
    """Map nama field backend ke label yang lebih mudah dibaca user."""
    labels = {
        'name': 'Nama lengkap',
        'nik': 'NIK',
        'address': 'Alamat',
        'phone': 'Nomor telepon',
        'emergency_phone': 'Kontak darurat',
        'date_of_birth': 'Tanggal lahir',
    }
    return labels.get(field_name, field_name)


def _build_booking_failure_response(response):
    """Normalisasi error booking agar chatbot selalu membalas dengan jelas."""
    payload = _safe_json_from_response(response)
    code = (payload.get('code') or '').strip().upper()
    message = payload.get('message') or payload.get('error') or 'Gagal membuat pesanan.'

    if code == 'PROFILE_INCOMPLETE':
        missing_fields = payload.get('missing_fields') or []
        missing_human = [_humanize_profile_field(field) for field in missing_fields]
        if missing_human:
            missing_text = ', '.join(missing_human)
            message = (
                "Data profil Anda belum lengkap, jadi booking belum bisa diproses. "
                f"Silakan lengkapi dulu: {missing_text}."
            )
        else:
            message = (
                "Data profil Anda belum lengkap, jadi booking belum bisa diproses. "
                "Silakan lengkapi data profil terlebih dahulu."
            )

        return {
            'success': False,
            'code': code,
            'next_step': payload.get('next_step', 'profile_screen'),
            'message': message,
        }

    if code == 'EXPERIENCE_ONBOARDING_REQUIRED':
        return {
            'success': False,
            'code': code,
            'next_step': payload.get('next_step', 'experience_onboarding'),
            'message': (
                "Booking belum bisa diproses karena data pengalaman mendaki belum diisi. "
                "Silakan selesaikan onboarding pengalaman pendakian terlebih dahulu."
            ),
        }

    if code == 'HIGH_RISK_CONFIRMATION_REQUIRED':
        return {
            'success': False,
            'code': code,
            'message': message,
            'dss': payload.get('dss'),
        }

    if response.status_code in (401, 403):
        fallback = (
            "Sesi login Anda tidak valid atau sudah berakhir. "
            "Silakan login ulang lalu coba booking lagi."
        )
        return {
            'success': False,
            'code': code or 'UNAUTHORIZED',
            'message': message if message and message != 'Unauthenticated.' else fallback,
        }

    if not message or message == 'Gagal membuat pesanan.':
        message = f"Gagal membuat pesanan (HTTP {response.status_code})."

    return {
        'success': False,
        'code': code or f'HTTP_{response.status_code}',
        'message': message,
    }


# ============================================
# TOOL FUNCTIONS (dipanggil oleh Gemini)
# ============================================

def tool_create_booking(
    user_id,
    id_gunung,
    id_jalur,
    tanggal_naik,
    tanggal_turun,
    anggota_ids=None,
    auth_token=None,
    force_continue=False,
):
    """Membuat booking + transaksi + pembayaran Midtrans sekaligus"""
    try:
        if not user_id or not id_gunung or not id_jalur or not tanggal_naik or not tanggal_turun:
            return {
                "success": False,
                "message": (
                    "Data booking belum lengkap. Pastikan user, gunung, jalur, "
                    "tanggal naik, dan tanggal turun sudah dipilih."
                ),
            }

        if not auth_token:
            return {
                "success": False,
                "code": "UNAUTHORIZED",
                "next_step": "login",
                "message": (
                    "Sesi login tidak ditemukan. Silakan login ulang, "
                    "lalu coba pemesanan lagi dari chatbot."
                ),
            }

        # Validasi format dan logika tanggal
        tanggal_naik_dt = datetime.datetime.strptime(tanggal_naik, '%Y-%m-%d')
        tanggal_turun_dt = datetime.datetime.strptime(tanggal_turun, '%Y-%m-%d')
        if tanggal_turun_dt < tanggal_naik_dt:
            return {
                "success": False,
                "message": (
                    "Tanggal turun tidak boleh sebelum tanggal naik. "
                    "Untuk tektok, gunakan tanggal naik dan turun yang sama."
                ),
            }
        
        # Ambil biaya dari jalur
        trails = fetch_trails_data()
        trail = next((t for t in trails if t['id'] == id_jalur), None)
        if not trail:
            return {"success": False, "message": f"Jalur dengan ID {id_jalur} tidak ditemukan"}
        
        total_harga = trail['biaya']
        nama_gunung = trail.get('nama_gunung', '-')
        nama_jalur = trail.get('nama_jalur', '-')

        # Normalisasi anggota tambahan (opsional)
        anggota_clean = _normalize_member_ids(anggota_ids, user_id=user_id)
        total_pendaki = 1 + len(anggota_clean)
        total_tagihan = total_harga * total_pendaki
        
        # Step 1: Buat booking/order
        response = requests.post(
            f"{LARAVEL_API_URL}/orders",
            json={
                "id_gunung": id_gunung,
                "id_jalur": id_jalur,
                "id_user": user_id,
                "tanggal_naik": tanggal_naik,
                "tanggal_turun": tanggal_turun,
                "total_harga_tiket": total_harga,
                "anggota_ids": anggota_clean,
                "force_continue": force_continue,
            },
            headers=_laravel_json_headers(auth_token),
            timeout=10
        )
        
        if response.status_code != 201:
            return _build_booking_failure_response(response)
        
        data = _safe_json_from_response(response)
        order_id = data.get('order', {}).get('id')
        
        if not order_id:
            return {"success": False, "message": "Pesanan dibuat tapi ID tidak ditemukan"}
        
        # Step 2: Buat transaksi
        try:
            tx_response = requests.post(
                f"{LARAVEL_API_URL}/transactions/store",
                json={"id_pesanan": order_id},
                headers=_laravel_json_headers(auth_token),
                timeout=10
            )
            
            if tx_response.status_code == 201:
                tx_data = _safe_json_from_response(tx_response)
                transaction_id = tx_data.get('transaction', {}).get('id')
                
                # Step 3: Buat pembayaran Midtrans
                if transaction_id:
                    try:
                        pay_response = requests.post(
                            f"{LARAVEL_API_URL}/midtrans/create-payment",
                            json={"order_id": order_id},
                            headers=_laravel_json_headers(auth_token),
                            timeout=10
                        )
                        
                        if pay_response.status_code in (200, 201):
                            pay_data = _safe_json_from_response(pay_response)
                            redirect_url = pay_data.get('data', {}).get('redirect_url', '')
                            
                            return {
                                "success": True,
                                "message": (f"Pemesanan berhasil dibuat!\n\n"
                                           f"Detail Pesanan:\n"
                                           f"- ID Pesanan: {order_id}\n"
                                           f"- Gunung: {nama_gunung}\n"
                                           f"- Jalur: {nama_jalur}\n"
                                           f"- Tanggal Naik: {tanggal_naik}\n"
                                           f"- Tanggal Turun: {tanggal_turun}\n"
                                           f"- Jumlah Pendaki: {total_pendaki} orang\n"
                                           f"- Biaya per Orang: Rp {total_harga:,}\n"
                                           f"- Total Biaya: Rp {total_tagihan:,}\n\n"
                                           f"Link pembayaran Midtrans sudah disiapkan. "
                                           f"Silakan klik tombol 'Bayar Sekarang' di bawah untuk melanjutkan pembayaran."),
                                "order_id": order_id,
                                "payment_url": redirect_url,
                                "transaction_id": transaction_id,
                            }
                        else:
                            print(f"Midtrans payment error: {pay_response.text}")
                    except Exception as e:
                        print(f"Error creating Midtrans payment: {e}")
                
                # Jika Midtrans gagal, masih berhasil booking
                return {
                    "success": True,
                    "message": (f"Pemesanan berhasil dibuat!\n\n"
                               f"Detail Pesanan:\n"
                               f"- ID Pesanan: {order_id}\n"
                               f"- Gunung: {nama_gunung}\n"
                               f"- Jalur: {nama_jalur}\n"
                               f"- Tanggal Naik: {tanggal_naik}\n"
                               f"- Tanggal Turun: {tanggal_turun}\n"
                               f"- Jumlah Pendaki: {total_pendaki} orang\n"
                               f"- Biaya per Orang: Rp {total_harga:,}\n"
                               f"- Total Biaya: Rp {total_tagihan:,}\n\n"
                               f"Silakan lanjutkan pembayaran melalui menu Transaksi di aplikasi."),
                    "order_id": order_id,
                }
            else:
                print(f"Transaction creation error: {tx_response.text}")
        except Exception as e:
            print(f"Error creating transaction: {e}")
        
        # Jika transaksi gagal, masih berhasil booking
        return {
            "success": True,
            "message": (f"Pemesanan berhasil dibuat! ID Pesanan: {order_id}\n"
                       f"Gunung: {nama_gunung}, Jalur: {nama_jalur}\n"
                       f"Tanggal: {tanggal_naik} - {tanggal_turun}\n"
                       f"Jumlah Pendaki: {total_pendaki} orang\n"
                       f"Biaya per Orang: Rp {total_harga:,}\n"
                       f"Total: Rp {total_tagihan:,}\n\n"
                       f"Silakan lanjutkan pembayaran melalui menu Transaksi di aplikasi."),
            "order_id": order_id,
        }
    except ValueError:
        return {
            "success": False,
            "message": (
                "Format tanggal tidak valid. Gunakan format YYYY-MM-DD "
                "(contoh: 2026-04-20)."
            ),
        }
    except Exception as e:
        print(f"Error membuat booking: {e}")
        return {
            "success": False,
            "message": "Maaf, terjadi kendala saat membuat booking. Silakan coba lagi.",
        }


def tool_get_sar_dashboard():
    """Mengambil data SAR dashboard"""
    panics = fetch_panic_data()
    active = [p for p in panics if p.get('status') in ('pending', 'active', 'in_progress')]
    
    result = {
        "total_permintaan": len(panics),
        "aktif": len(active),
        "data_aktif": []
    }
    
    for p in active:
        result["data_aktif"].append({
            "id": p['id'],
            "user": p.get('nama_user', '-'),
            "telepon": p.get('telepon_user', '-'),
            "darurat": p.get('emergency_phone', '-'),
            "gunung": p.get('nama_gunung', '-'),
            "jalur": p.get('nama_jalur', '-'),
            "tipe_darurat": p['emergency_type'],
            "deskripsi": p.get('description', '-'),
            "koordinat": f"{p.get('latitude', '-')}, {p.get('longitude', '-')}",
            "status": p['status'],
            "waktu": str(p['created_at'])
        })
    
    return result


def tool_export_excel(data_type, role, user_id=None):
    """Generate file Excel berdasarkan tipe data yang diminta"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        filename = ""
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        
        guard_trail_ids = None
        if role == 'penjaga' and user_id:
            guard_trails = fetch_trails_by_guard(user_id)
            guard_trail_ids = [t['id'] for t in guard_trails] if guard_trails else []

        if data_type == "sar_dashboard":
            if role == 'penjaga' and user_id:
                panics = fetch_panics_by_trail_ids(guard_trail_ids)
            else:
                panics = fetch_panic_data()
            ws.title = "SAR Dashboard"
            headers = ["ID", "Nama User", "Telepon", "Telepon Darurat", "Gunung", "Jalur", 
                       "Tipe Darurat", "Deskripsi", "Latitude", "Longitude", "Status", "Waktu"]
            ws.append(headers)
            for p in panics:
                ws.append([
                    p['id'], p.get('nama_user', '-'), p.get('telepon_user', '-'),
                    p.get('emergency_phone', '-'), p.get('nama_gunung', '-'),
                    p.get('nama_jalur', '-'), p['emergency_type'],
                    p.get('description', '-'), p.get('latitude', '-'),
                    p.get('longitude', '-'), p['status'], str(p['created_at'])
                ])
            filename = f"rekap_sar_dashboard_{timestamp}.xlsx"
            
        elif data_type == "laporan_pendapatan":
            if role == 'penjaga' and user_id:
                transactions = fetch_transactions_by_trail_ids(guard_trail_ids)
            else:
                transactions = fetch_transactions_data()
            ws.title = "Laporan Pendapatan"
            headers = ["ID Transaksi", "ID Pesanan", "Nama User", "Gunung", "Jalur",
                       "Total Bayar", "Status", "Tipe Pembayaran", "Waktu Pembayaran", "Dibuat"]
            ws.append(headers)
            total_pendapatan = 0
            for t in transactions:
                status_clean = str(t.get('status_pesanan', '')).strip().lower()
                if status_clean in ('complete', 'completed', 'paid', 'confirmed', 'success', 'settlement'):
                    ws.append([
                        t['id'], t['id_pesanan'], t.get('nama_user', '-'),
                        t.get('nama_gunung', '-'), t.get('nama_jalur', '-'),
                        t['total_bayar'], t['status_pesanan'],
                        t.get('payment_type', '-'), str(t.get('waktu_pembayaran', '-')),
                        str(t['created_at'])
                    ])
                    try:
                        total_pendapatan += float(t.get('total_bayar', 0) or 0)
                    except (ValueError, TypeError):
                        pass
            
            if isinstance(total_pendapatan, float) and total_pendapatan.is_integer():
                total_pendapatan = int(total_pendapatan)
            
            # Add total row
            ws.append([])
            ws.append(["", "", "", "", "TOTAL PENDAPATAN:", total_pendapatan])
            filename = f"rekap_laporan_pendapatan_{timestamp}.xlsx"
            
        elif data_type == "pesanan":
            if role == 'penjaga' and user_id:
                orders = fetch_orders_by_trail_ids(guard_trail_ids)
            else:
                orders = fetch_orders_data()
            ws.title = "Data Pesanan"
            headers = ["ID", "Nama User", "Email", "Gunung", "Jalur",
                       "Tanggal Naik", "Tanggal Turun", "Total Harga", "Status", "Dibuat"]
            ws.append(headers)
            for o in orders:
                ws.append([
                    o['id'], o.get('nama_user', '-'), o.get('email_user', '-'),
                    o.get('nama_gunung', '-'), o.get('nama_jalur', '-'),
                    str(o['tanggal_naik']), str(o['tanggal_turun']),
                    o['total_harga_tiket'], o['status'], str(o['created_at'])
                ])
            filename = f"rekap_pesanan_{timestamp}.xlsx"
            
        elif data_type == "transaksi":
            if role == 'penjaga' and user_id:
                transactions = fetch_transactions_by_trail_ids(guard_trail_ids)
            else:
                transactions = fetch_transactions_data()
            ws.title = "Data Transaksi"
            headers = ["ID", "ID Pesanan", "Nama User", "Gunung", "Jalur",
                       "Total Bayar", "Status", "Tipe Pembayaran", "Waktu Pembayaran"]
            ws.append(headers)
            for t in transactions:
                ws.append([
                    t['id'], t['id_pesanan'], t.get('nama_user', '-'),
                    t.get('nama_gunung', '-'), t.get('nama_jalur', '-'),
                    t['total_bayar'], t['status_pesanan'],
                    t.get('payment_type', '-'), str(t.get('waktu_pembayaran', '-'))
                ])
            filename = f"rekap_transaksi_{timestamp}.xlsx"
            
        elif data_type == "user":
            users = fetch_users_data()
            ws.title = "Data User"
            headers = ["ID", "Nama", "Email", "Telepon", "Alamat", "NIK", 
                       "Telepon Darurat", "Tanggal Lahir", "Level", "Terdaftar"]
            ws.append(headers)
            for u in users:
                ws.append([
                    u['id'], u['name'], u['email'], u.get('phone', '-'),
                    u.get('address', '-'), u.get('nik', '-'),
                    u.get('emergency_phone', '-'), str(u.get('date_of_birth', '-')),
                    u.get('level', '-'), str(u['created_at'])
                ])
            filename = f"rekap_data_user_{timestamp}.xlsx"
            
        elif data_type == "gunung":
            mountains = fetch_mountains_data()
            ws.title = "Data Gunung"
            headers = ["ID", "Nama", "Ketinggian (mdpl)", "Provinsi", "Latitude", "Longitude"]
            ws.append(headers)
            for m in mountains:
                ws.append([
                    m['id'], m['nama'], m['ketinggian'], m.get('provinsi', '-'),
                    m.get('latitude', '-'), m.get('longitude', '-')
                ])
            filename = f"rekap_data_gunung_{timestamp}.xlsx"
            
        elif data_type == "jalur":
            trails = fetch_trails_data()
            ws.title = "Data Jalur"
            headers = ["ID", "Nama Jalur", "Gunung", "Jarak (km)", "Biaya (Rp)",
                       "Provinsi", "Kabupaten", "Kecamatan", "Desa"]
            ws.append(headers)
            for t in trails:
                ws.append([
                    t['id'], t['nama_jalur'], t['nama_gunung'], t['jarak'],
                    t['biaya'], t.get('provinsi', '-'), t.get('kabupaten', '-'),
                    t.get('kecamatan', '-'), t.get('desa', '-')
                ])
            filename = f"rekap_data_jalur_{timestamp}.xlsx"
        else:
            return {"success": False, "message": f"Tipe data '{data_type}' tidak dikenali"}
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        filepath = os.path.join(EXPORT_DIR, filename)
        wb.save(filepath)
        
        return {
            "success": True,
            "message": f"File Excel berhasil dibuat: {filename}",
            "filename": filename,
            "filepath": filepath
        }
    except ImportError:
        return {"success": False, "message": "Library openpyxl belum terinstall. Jalankan: pip install openpyxl"}
    except Exception as e:
        return {"success": False, "message": f"Error membuat Excel: {str(e)}"}


def tool_crud_mountain(action, data=None):
    """CRUD operasi untuk gunung via API Laravel"""
    try:
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                pass
        
        if not isinstance(data, dict):
            data = {}

        if action == "list":
            mountains = fetch_mountains_data()
            return {"success": True, "data": mountains, "total": len(mountains)}
        
        elif action == "create" and data:
            name_val = data.get('nama') or data.get('nama_gunung')
            name_str = f"Gunung {name_val}" if name_val else "Gunung baru"
            response = requests.post(
                f"{LARAVEL_API_URL}/mountains",
                json=data,
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code in (200, 201):
                return {"success": True, "message": f"{name_str} berhasil ditambahkan.", "data": response.json()}
            else:
                return {"success": False, "message": f"Gagal menambahkan data: {response.json().get('message', response.text)}"}
        
        elif action == "update" and data:
            mountain_id = data.pop('id', None)
            mountain_name = data.get('nama') or data.get('nama_gunung')
            if not mountain_id:
                return {"success": False, "message": "Data gunung tidak ditemukan untuk diperbarui."}
            
            if not mountain_name:
                mountains = fetch_mountains_data()
                m = next((item for item in mountains if str(item['id']) == str(mountain_id)), None)
                if m:
                    mountain_name = m.get('nama')
            
            name_str = f"Gunung {mountain_name}" if mountain_name else "Data gunung"

            response = requests.put(
                f"{LARAVEL_API_URL}/mountains/{mountain_id}",
                json=data,
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code == 200:
                return {"success": True, "message": f"{name_str} berhasil diperbarui."}
            else:
                return {"success": False, "message": f"Gagal memperbarui data: {response.json().get('message', response.text)}"}
        
        elif action == "delete" and data:
            mountain_id = data.get('id')
            if not mountain_id:
                return {"success": False, "message": "Data gunung tidak ditemukan untuk dihapus."}
            
            mountains = fetch_mountains_data()
            m = next((item for item in mountains if str(item['id']) == str(mountain_id)), None)
            mountain_name = m.get('nama') if m else None
            name_str = f"Gunung {mountain_name}" if mountain_name else "Data gunung"

            response = requests.delete(
                f"{LARAVEL_API_URL}/mountains/{mountain_id}",
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code in (200, 204):
                return {"success": True, "message": f"{name_str} berhasil dihapus."}
            else:
                return {"success": False, "message": f"Gagal menghapus data: {response.text}"}
        
        return {"success": False, "message": f"Aksi '{action}' tidak valid."}
    except Exception as e:
        return {"success": False, "message": f"Terjadi kesalahan saat memproses data gunung: {str(e)}"}


def tool_crud_trail(action, data=None):
    """CRUD operasi untuk jalur via API Laravel"""
    try:
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                pass
        
        if not isinstance(data, dict):
            data = {}

        if action == "list":
            trails = fetch_trails_data()
            return {"success": True, "data": trails, "total": len(trails)}
        
        elif action == "create" and data:
            name_val = data.get('nama') or data.get('nama_jalur')
            name_str = f"Jalur {name_val}" if name_val else "Jalur baru"
            response = requests.post(
                f"{LARAVEL_API_URL}/routes",
                json=data,
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code in (200, 201):
                return {"success": True, "message": f"{name_str} berhasil ditambahkan.", "data": response.json()}
            else:
                return {"success": False, "message": f"Gagal menambahkan data: {response.json().get('message', response.text)}"}
        
        elif action == "update" and data:
            trail_id = data.pop('id', None)
            trail_name = data.get('nama') or data.get('nama_jalur')
            if not trail_id:
                return {"success": False, "message": "Data jalur tidak ditemukan untuk diperbarui."}
            
            if not trail_name:
                trails = fetch_trails_data()
                t = next((item for item in trails if str(item['id']) == str(trail_id)), None)
                if t:
                    trail_name = t.get('nama_jalur') or t.get('nama')
            
            name_str = f"Jalur {trail_name}" if trail_name else "Data jalur"

            response = requests.put(
                f"{LARAVEL_API_URL}/routes/{trail_id}",
                json=data,
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code == 200:
                return {"success": True, "message": f"{name_str} berhasil diperbarui."}
            else:
                return {"success": False, "message": f"Gagal memperbarui data: {response.json().get('message', response.text)}"}
        
        elif action == "delete" and data:
            trail_id = data.get('id')
            if not trail_id:
                return {"success": False, "message": "Data jalur tidak ditemukan untuk dihapus."}
            
            trails = fetch_trails_data()
            t = next((item for item in trails if str(item['id']) == str(trail_id)), None)
            trail_name = t.get('nama_jalur') or t.get('nama') if t else None
            name_str = f"Jalur {trail_name}" if trail_name else "Data jalur"

            response = requests.delete(
                f"{LARAVEL_API_URL}/routes/{trail_id}",
                headers=_laravel_json_headers(chatbot_crud=True),
                timeout=10
            )
            if response.status_code in (200, 204):
                return {"success": True, "message": f"{name_str} berhasil dihapus."}
            else:
                return {"success": False, "message": f"Gagal menghapus data: {response.text}"}
        
        return {"success": False, "message": f"Aksi '{action}' tidak valid."}
    except Exception as e:
        return {"success": False, "message": f"Terjadi kesalahan saat memproses data jalur: {str(e)}"}
