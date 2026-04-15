"""
My Hiking Chatbot Backend
=========================
Chatbot menggunakan Gemini API dengan RAG (Retrieval Augmented Generation)
untuk menjawab pertanyaan seputar pendakian gunung.

Mendukung 3 role:
- pendaki: informasi gunung + pemesanan tiket via chatbot
- admin: CRUD data + ekspor Excel
- penjaga: SAR dashboard + ekspor Excel

Author: My Hiking Team
Date: March 2024
"""

import os
import json
import uuid
import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from dotenv import load_dotenv
import google.generativeai as genai
import pymysql
import requests

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
CORS(app)  # Enable CORS for Flutter app

# Configure Gemini API
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if GEMINI_API_KEY and GEMINI_API_KEY != 'your_gemini_api_key_here':
    genai.configure(api_key=GEMINI_API_KEY)

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', '127.0.0.1'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'database': os.getenv('DB_NAME') or os.getenv('DB_DATABASE', 'myhiking'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# Laravel API base URL
LARAVEL_API_URL = os.getenv('LARAVEL_API_URL', 'http://127.0.0.1:8000/api')

# Directory untuk menyimpan file Excel yang di-generate
EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)

# Initialize last_payment_url attribute  
get_gemini_response_last_payment_url = None


def get_db_connection():
    """Membuat koneksi ke database MySQL"""
    return pymysql.connect(**DB_CONFIG)


def clean_markdown(text):
    """Membersihkan format markdown dari teks respons Gemini"""
    import re
    # Hapus bold (**text** atau __text__)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    # Hapus italic (*text* atau _text_) - hati-hati dengan bullet points
    text = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'\1', text)
    # Hapus heading markers (# ## ### etc)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    # Hapus backtick code blocks
    text = re.sub(r'```[\s\S]*?```', '', text)
    text = re.sub(r'`(.+?)`', r'\1', text)
    # Bersihkan multiple newlines
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def init_chat_history_table():
    """Membuat tabel chat_histories jika belum ada"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS chat_histories (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    role VARCHAR(20) NOT NULL DEFAULT 'pendaki',
                    title VARCHAR(255) DEFAULT NULL,
                    messages LONGTEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_user_role (user_id, role)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
        conn.commit()
        conn.close()
        print("[OK] Tabel chat_histories siap")
    except Exception as e:
        print(f"[Warning] Gagal membuat tabel chat_histories: {e}")


# Inisialisasi tabel chat history saat startup
init_chat_history_table()


# ============================================
# DATABASE FETCH FUNCTIONS
# ============================================

def fetch_mountains_data():
    """Mengambil data gunung dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT m.id, m.nama, m.deskripsi, m.ketinggian, m.latitude, m.longitude,
                       p.name as provinsi, r.name as kabupaten, d.name as kecamatan, v.name as desa
                FROM mountains m
                LEFT JOIN reg_provinces p ON m.province_id = p.id
                LEFT JOIN reg_regencies r ON m.regency_id = r.id
                LEFT JOIN reg_districts d ON m.district_id = d.id
                LEFT JOIN reg_villages v ON m.village_id = v.id
            """
            cursor.execute(query)
            mountains = cursor.fetchall()
        conn.close()
        return mountains
    except Exception as e:
        print(f"Error fetching mountains: {e}")
        return []


def fetch_trails_data():
    """Mengambil data jalur pendakian dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT r.id, r.nama as nama_jalur, r.jarak, r.deskripsi, r.biaya, r.latitude, r.longitude,
                       m.nama as nama_gunung, m.id as id_gunung, m.ketinggian,
                       p.name as provinsi, rg.name as kabupaten, d.name as kecamatan, v.name as desa
                FROM routes r
                LEFT JOIN mountains m ON r.id_gunung = m.id
                LEFT JOIN reg_provinces p ON r.province_id = p.id
                LEFT JOIN reg_regencies rg ON r.regency_id = rg.id
                LEFT JOIN reg_districts d ON r.district_id = d.id
                LEFT JOIN reg_villages v ON r.village_id = v.id
            """
            cursor.execute(query)
            trails = cursor.fetchall()
        conn.close()
        return trails
    except Exception as e:
        print(f"Error fetching trails: {e}")
        return []


def fetch_rules_data():
    """Mengambil data tata tertib dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT ru.id, ru.description as tata_tertib, 
                       r.nama as nama_jalur, m.nama as nama_gunung
                FROM rules ru
                LEFT JOIN routes r ON ru.jalur_id = r.id
                LEFT JOIN mountains m ON r.id_gunung = m.id
            """
            cursor.execute(query)
            rules = cursor.fetchall()
        conn.close()
        return rules
    except Exception as e:
        print(f"Error fetching rules: {e}")
        return []


def fetch_orders_data():
    """Mengambil data pesanan dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT o.id, o.id_gunung, o.id_jalur, o.id_user, 
                       o.tanggal_naik, o.tanggal_turun, o.total_harga_tiket, o.status,
                       o.created_at, o.updated_at,
                       m.nama as nama_gunung, r.nama as nama_jalur,
                       u.name as nama_user, u.email as email_user
                FROM orders o
                LEFT JOIN mountains m ON o.id_gunung = m.id
                LEFT JOIN routes r ON o.id_jalur = r.id
                LEFT JOIN users u ON o.id_user = u.id
                ORDER BY o.created_at DESC
            """
            cursor.execute(query)
            orders = cursor.fetchall()
        conn.close()
        return orders
    except Exception as e:
        print(f"Error fetching orders: {e}")
        return []


def fetch_transactions_data():
    """Mengambil data transaksi dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT t.id, t.id_pesanan, t.total_bayar, t.status_pesanan,
                       t.waktu_pembayaran, t.bukti, t.payment_type,
                       t.created_at, t.updated_at,
                       o.id_gunung, o.id_jalur, o.id_user, o.tanggal_naik,
                       m.nama as nama_gunung, r.nama as nama_jalur,
                       u.name as nama_user
                FROM transactions t
                LEFT JOIN orders o ON t.id_pesanan = o.id
                LEFT JOIN mountains m ON o.id_gunung = m.id
                LEFT JOIN routes r ON o.id_jalur = r.id
                LEFT JOIN users u ON o.id_user = u.id
                ORDER BY t.created_at DESC
            """
            cursor.execute(query)
            transactions = cursor.fetchall()
        conn.close()
        return transactions
    except Exception as e:
        print(f"Error fetching transactions: {e}")
        return []


def fetch_panic_data():
    """Mengambil data panic/SAR dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT p.id, p.user_id, p.order_id, p.latitude, p.longitude,
                       p.emergency_type, p.description, p.status,
                       p.created_at, p.updated_at,
                       u.name as nama_user, u.phone as telepon_user, u.emergency_phone,
                       o.id_gunung, o.id_jalur, o.tanggal_naik,
                       m.nama as nama_gunung, r.nama as nama_jalur
                FROM panic_requests p
                LEFT JOIN users u ON p.user_id = u.id
                LEFT JOIN orders o ON p.order_id = o.id
                LEFT JOIN mountains m ON o.id_gunung = m.id
                LEFT JOIN routes r ON o.id_jalur = r.id
                ORDER BY p.created_at DESC
            """
            cursor.execute(query)
            panics = cursor.fetchall()
        conn.close()
        return panics
    except Exception as e:
        print(f"Error fetching panics: {e}")
        return []


def fetch_users_data():
    """Mengambil data user dari database"""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT u.id, u.name, u.email, u.phone, u.address, u.nik,
                       u.emergency_phone, u.date_of_birth, u.level,
                       u.created_at, u.updated_at
                FROM users u
                ORDER BY u.created_at DESC
            """
            cursor.execute(query)
            users = cursor.fetchall()
        conn.close()
        return users
    except Exception as e:
        print(f"Error fetching users: {e}")
        return []


# ============================================
# CONTEXT BUILDERS
# ============================================

def build_context_pendaki():
    """Membangun konteks untuk chatbot pendaki"""
    context_parts = []
    
    mountains = fetch_mountains_data()
    trails = fetch_trails_data()
    rules = fetch_rules_data()
    mountains_sorted = []
    trails_sorted = []
    
    # Build mountain context (user-facing: nomor urut, tanpa ID mentah)
    if mountains:
        context_parts.append("=== DATA GUNUNG ===")
        mountains_sorted = sorted(mountains, key=lambda x: x.get('id', 0))
        for idx, m in enumerate(mountains_sorted, start=1):
            mountain_info = f"""
No Gunung: {idx}
Nama Gunung: {m['nama']}
Ketinggian: {m['ketinggian']} mdpl
Lokasi: {m.get('desa', '')}, {m.get('kecamatan', '')}, {m.get('kabupaten', '')}, {m.get('provinsi', '')}
Deskripsi: {m['deskripsi']}
Koordinat: {m.get('latitude', '-')}, {m.get('longitude', '-')}
---"""
            context_parts.append(mountain_info)
    
    # Build trails context (user-facing: nomor urut, tanpa ID mentah)
    if trails:
        context_parts.append("\n=== DATA JALUR PENDAKIAN ===")
        trails_sorted = sorted(trails, key=lambda x: x.get('id', 0))
        for idx, t in enumerate(trails_sorted, start=1):
            trail_info = f"""
No Jalur: {idx}
Nama Jalur: {t['nama_jalur']}
Gunung: {t['nama_gunung']} ({t.get('ketinggian', '-')} mdpl)
Jarak: {t['jarak']} km
Biaya: Rp {t['biaya']:,}
Lokasi Basecamp: {t.get('desa', '')}, {t.get('kecamatan', '')}, {t.get('kabupaten', '')}, {t.get('provinsi', '')}
Deskripsi: {t.get('deskripsi', '-')}
---"""
            context_parts.append(trail_info)

        # Mapping internal untuk function calling (jangan ditampilkan ke user)
        context_parts.append("\n=== INTERNAL_ONLY_MAPPINGS (JANGAN DITAMPILKAN KE USER) ===")
        for idx, m in enumerate(mountains_sorted, start=1):
            context_parts.append(f"MAP_GUNUNG: nomor {idx} -> id_gunung {m['id']} ({m['nama']})")
        for idx, t in enumerate(trails_sorted, start=1):
            context_parts.append(f"MAP_JALUR: nomor {idx} -> id_jalur {t['id']} ({t['nama_jalur']})")
    
    # Build rules context
    if rules:
        context_parts.append("\n=== DATA TATA TERTIB ===")
        for r in rules:
            rule_info = f"""
Jalur: {r['nama_jalur']} - Gunung {r['nama_gunung']}
Tata Tertib: {r['tata_tertib']}
---"""
            context_parts.append(rule_info)
    
    return "\n".join(context_parts)


def build_context_admin():
    """Membangun konteks untuk chatbot admin"""
    context_parts = []
    
    mountains = fetch_mountains_data()
    trails = fetch_trails_data()
    rules = fetch_rules_data()
    orders = fetch_orders_data()
    transactions = fetch_transactions_data()
    users = fetch_users_data()
    panics = fetch_panic_data()
    
    # Mountains
    if mountains:
        context_parts.append("=== DATA GUNUNG ===")
        for m in mountains:
            context_parts.append(f"ID: {m['id']} | Nama: {m['nama']} | Ketinggian: {m['ketinggian']} mdpl | Lokasi: {m.get('kabupaten', '')}, {m.get('provinsi', '')}")
    
    # Trails
    if trails:
        context_parts.append("\n=== DATA JALUR PENDAKIAN ===")
        for t in trails:
            context_parts.append(f"ID: {t['id']} | Jalur: {t['nama_jalur']} | Gunung: {t['nama_gunung']} (ID: {t['id_gunung']}) | Biaya: Rp {t['biaya']:,} | Jarak: {t['jarak']} km")
    
    # Orders (recent 50)
    if orders:
        context_parts.append(f"\n=== DATA PESANAN (Total: {len(orders)}, Ditampilkan 50 terbaru) ===")
        for o in orders[:50]:
            context_parts.append(f"ID: {o['id']} | User: {o.get('nama_user', '-')} | Gunung: {o.get('nama_gunung', '-')} | Jalur: {o.get('nama_jalur', '-')} | Tanggal Naik: {o['tanggal_naik']} | Status: {o['status']} | Total: Rp {o['total_harga_tiket']:,}")
    
    # Transactions (recent 50)
    if transactions:
        context_parts.append(f"\n=== DATA TRANSAKSI (Total: {len(transactions)}, Ditampilkan 50 terbaru) ===")
        for t in transactions[:50]:
            context_parts.append(f"ID: {t['id']} | Pesanan: {t['id_pesanan']} | User: {t.get('nama_user', '-')} | Total: Rp {t['total_bayar']:,} | Status: {t['status_pesanan']} | Pembayaran: {t.get('waktu_pembayaran', '-')}")
    
    # Users
    if users:
        context_parts.append(f"\n=== DATA USER (Total: {len(users)}) ===")
        for u in users[:50]:
            context_parts.append(f"ID: {u['id']} | Nama: {u['name']} | Email: {u['email']} | Phone: {u.get('phone', '-')} | Level: {u.get('level', '-')}")
    
    # Panics/SAR
    if panics:
        context_parts.append(f"\n=== DATA SAR/DARURAT (Total: {len(panics)}) ===")
        for p in panics:
            context_parts.append(f"ID: {p['id']} | User: {p.get('nama_user', '-')} | Gunung: {p.get('nama_gunung', '-')} | Jalur: {p.get('nama_jalur', '-')} | Tipe: {p['emergency_type']} | Status: {p['status']} | Waktu: {p['created_at']}")
    
    return "\n".join(context_parts)


def build_context_penjaga():
    """Membangun konteks untuk chatbot penjaga jalur"""
    context_parts = []
    
    orders = fetch_orders_data()
    panics = fetch_panic_data()
    trails = fetch_trails_data()
    mountains = fetch_mountains_data()
    
    # Active panics (SAR Dashboard)
    active_panics = [p for p in panics if p.get('status') in ('pending', 'active', 'in_progress')]
    if active_panics:
        context_parts.append("⚠️ === PERMINTAAN SAR AKTIF (PERLU PERHATIAN!) ===")
        for p in active_panics:
            context_parts.append(f"🚨 ID: {p['id']} | User: {p.get('nama_user', '-')} | Telepon: {p.get('telepon_user', '-')} | Darurat: {p.get('emergency_phone', '-')} | Gunung: {p.get('nama_gunung', '-')} | Jalur: {p.get('nama_jalur', '-')} | Tipe: {p['emergency_type']} | Deskripsi: {p.get('description', '-')} | Koordinat: {p.get('latitude', '-')}, {p.get('longitude', '-')} | Status: {p['status']} | Waktu: {p['created_at']}")
    else:
        context_parts.append("✅ Tidak ada permintaan SAR aktif saat ini.")
    
    # All panics history
    if panics:
        context_parts.append(f"\n=== RIWAYAT SAR/DARURAT (Total: {len(panics)}) ===")
        for p in panics[:30]:
            context_parts.append(f"ID: {p['id']} | User: {p.get('nama_user', '-')} | Gunung: {p.get('nama_gunung', '-')} | Tipe: {p['emergency_type']} | Status: {p['status']} | Waktu: {p['created_at']}")
    
    # Active orders (pendaki yang sedang di gunung) 
    if orders:
        active_orders = [o for o in orders if o.get('status') in ('confirmed', 'active', 'checked_in')]
        if active_orders:
            context_parts.append(f"\n=== PESANAN AKTIF / PENDAKI DI GUNUNG (Total: {len(active_orders)}) ===")
            for o in active_orders[:30]:
                context_parts.append(f"ID: {o['id']} | User: {o.get('nama_user', '-')} | Gunung: {o.get('nama_gunung', '-')} | Jalur: {o.get('nama_jalur', '-')} | Tanggal Naik: {o['tanggal_naik']} | Tanggal Turun: {o['tanggal_turun']} | Status: {o['status']}")
        
        context_parts.append(f"\n=== SEMUA PESANAN (Total: {len(orders)}, 50 terbaru) ===")
        for o in orders[:50]:
            context_parts.append(f"ID: {o['id']} | User: {o.get('nama_user', '-')} | Gunung: {o.get('nama_gunung', '-')} | Status: {o['status']} | Tanggal: {o['tanggal_naik']}")
    
    # Mountains & Trails for reference
    if mountains:
        context_parts.append("\n=== DATA GUNUNG ===")
        for m in mountains:
            context_parts.append(f"ID: {m['id']} | {m['nama']} | {m['ketinggian']} mdpl")
    
    if trails:
        context_parts.append("\n=== DATA JALUR ===")
        for t in trails:
            context_parts.append(f"ID: {t['id']} | {t['nama_jalur']} | Gunung: {t['nama_gunung']}")
    
    return "\n".join(context_parts)


# ============================================
# TOOL FUNCTIONS (dipanggil oleh Gemini)
# ============================================

def _normalize_member_ids(raw_member_ids, user_id=None):
    """Normalisasi data anggota agar selalu list[int] unik dan valid."""
    if raw_member_ids is None:
        return []

    if isinstance(raw_member_ids, (int, str)):
        raw_member_ids = [raw_member_ids]
    elif not isinstance(raw_member_ids, list):
        return []

    normalized = []
    for uid in raw_member_ids:
        try:
            parsed_uid = int(uid)
            if parsed_uid <= 0:
                continue
            if user_id is not None and parsed_uid == int(user_id):
                continue
            normalized.append(parsed_uid)
        except Exception:
            continue

    return list(dict.fromkeys(normalized))


def _normalize_positive_int(raw_value):
    """Konversi nilai numerik ke int positif (mendukung 1, "1", 1.0)."""
    if raw_value is None or isinstance(raw_value, bool):
        return None

    try:
        if isinstance(raw_value, int):
            parsed = raw_value
        elif isinstance(raw_value, float):
            if not raw_value.is_integer():
                return None
            parsed = int(raw_value)
        elif isinstance(raw_value, str):
            text = raw_value.strip()
            if not text:
                return None
            if '.' in text:
                as_float = float(text)
                if not as_float.is_integer():
                    return None
                parsed = int(as_float)
            else:
                parsed = int(text)
        else:
            parsed = int(raw_value)
    except Exception:
        return None

    return parsed if parsed > 0 else None


def _laravel_json_headers(auth_token=None):
    """Bangun header JSON + Bearer token (jika tersedia)."""
    headers = {'Content-Type': 'application/json'}
    token = (auth_token or '').strip()
    if token.lower().startswith('bearer '):
        token = token[7:].strip()
    if token:
        headers['Authorization'] = f'Bearer {token}'
    return headers


def _safe_json_from_response(response):
    """Ambil JSON response secara aman tanpa melempar exception."""
    try:
        return response.json()
    except Exception:
        return {}


def _humanize_profile_field(field_name):
    """Map nama field backend ke label yang lebih mudah dibaca user."""
    labels = {
        'name': 'Nama lengkap',
        'nik': 'NIK',
        'address': 'Alamat',
        'phone': 'Nomor telepon',
        'emergency_phone': 'Kontak darurat',
        'date_of_birth': 'Tanggal lahir',
    }
    return labels.get(field_name, field_name)


def _build_booking_failure_response(response):
    """Normalisasi error booking agar chatbot selalu membalas dengan jelas."""
    payload = _safe_json_from_response(response)
    code = (payload.get('code') or '').strip().upper()
    message = payload.get('message') or payload.get('error') or 'Gagal membuat pesanan.'

    if code == 'PROFILE_INCOMPLETE':
        missing_fields = payload.get('missing_fields') or []
        missing_human = [_humanize_profile_field(field) for field in missing_fields]
        if missing_human:
            missing_text = ', '.join(missing_human)
            message = (
                "Data profil Anda belum lengkap, jadi booking belum bisa diproses. "
                f"Silakan lengkapi dulu: {missing_text}."
            )
        else:
            message = (
                "Data profil Anda belum lengkap, jadi booking belum bisa diproses. "
                "Silakan lengkapi data profil terlebih dahulu."
            )

        return {
            'success': False,
            'code': code,
            'next_step': payload.get('next_step', 'profile_screen'),
            'message': message,
        }

    if code == 'EXPERIENCE_ONBOARDING_REQUIRED':
        return {
            'success': False,
            'code': code,
            'next_step': payload.get('next_step', 'experience_onboarding'),
            'message': (
                "Booking belum bisa diproses karena data pengalaman mendaki belum diisi. "
                "Silakan selesaikan onboarding pengalaman pendakian terlebih dahulu."
            ),
        }

    if code == 'HIGH_RISK_CONFIRMATION_REQUIRED':
        return {
            'success': False,
            'code': code,
            'message': message,
            'dss': payload.get('dss'),
        }

    if response.status_code in (401, 403):
        fallback = (
            "Sesi login Anda tidak valid atau sudah berakhir. "
            "Silakan login ulang lalu coba booking lagi."
        )
        return {
            'success': False,
            'code': code or 'UNAUTHORIZED',
            'message': message if message and message != 'Unauthenticated.' else fallback,
        }

    if not message or message == 'Gagal membuat pesanan.':
        message = f"Gagal membuat pesanan (HTTP {response.status_code})."

    return {
        'success': False,
        'code': code or f'HTTP_{response.status_code}',
        'message': message,
    }


def tool_create_booking(
    user_id,
    id_gunung,
    id_jalur,
    tanggal_naik,
    tanggal_turun,
    anggota_ids=None,
    auth_token=None,
):
    """Membuat booking + transaksi + pembayaran Midtrans sekaligus"""
    try:
        if not user_id or not id_gunung or not id_jalur or not tanggal_naik or not tanggal_turun:
            return {
                "success": False,
                "message": (
                    "Data booking belum lengkap. Pastikan user, gunung, jalur, "
                    "tanggal naik, dan tanggal turun sudah dipilih."
                ),
            }

        if not auth_token:
            return {
                "success": False,
                "code": "UNAUTHORIZED",
                "next_step": "login",
                "message": (
                    "Sesi login tidak ditemukan. Silakan login ulang, "
                    "lalu coba pemesanan lagi dari chatbot."
                ),
            }

        # Validasi format dan logika tanggal
        tanggal_naik_dt = datetime.datetime.strptime(tanggal_naik, '%Y-%m-%d')
        tanggal_turun_dt = datetime.datetime.strptime(tanggal_turun, '%Y-%m-%d')
        if tanggal_turun_dt < tanggal_naik_dt:
            return {
                "success": False,
                "message": (
                    "Tanggal turun tidak boleh sebelum tanggal naik. "
                    "Untuk tektok, gunakan tanggal naik dan turun yang sama."
                ),
            }
        
        # Ambil biaya dari jalur
        trails = fetch_trails_data()
        trail = next((t for t in trails if t['id'] == id_jalur), None)
        if not trail:
            return {"success": False, "message": f"Jalur dengan ID {id_jalur} tidak ditemukan"}
        
        total_harga = trail['biaya']
        nama_gunung = trail.get('nama_gunung', '-')
        nama_jalur = trail.get('nama_jalur', '-')

        # Normalisasi anggota tambahan (opsional)
        anggota_clean = _normalize_member_ids(anggota_ids, user_id=user_id)
        total_pendaki = 1 + len(anggota_clean)
        total_tagihan = total_harga * total_pendaki
        
        # Step 1: Buat booking/order
        response = requests.post(
            f"{LARAVEL_API_URL}/orders",
            json={
                "id_gunung": id_gunung,
                "id_jalur": id_jalur,
                "id_user": user_id,
                "tanggal_naik": tanggal_naik,
                "tanggal_turun": tanggal_turun,
                "total_harga_tiket": total_harga,
                "anggota_ids": anggota_clean,
            },
            headers=_laravel_json_headers(auth_token),
            timeout=10
        )
        
        if response.status_code != 201:
            return _build_booking_failure_response(response)
        
        data = _safe_json_from_response(response)
        order_id = data.get('order', {}).get('id')
        
        if not order_id:
            return {"success": False, "message": "Pesanan dibuat tapi ID tidak ditemukan"}
        
        # Step 2: Buat transaksi
        try:
            tx_response = requests.post(
                f"{LARAVEL_API_URL}/transactions/store",
                json={"id_pesanan": order_id},
                headers=_laravel_json_headers(auth_token),
                timeout=10
            )
            
            if tx_response.status_code == 201:
                tx_data = _safe_json_from_response(tx_response)
                transaction_id = tx_data.get('transaction', {}).get('id')
                
                # Step 3: Buat pembayaran Midtrans
                if transaction_id:
                    try:
                        pay_response = requests.post(
                            f"{LARAVEL_API_URL}/midtrans/create-payment",
                            json={"order_id": order_id},
                            headers=_laravel_json_headers(auth_token),
                            timeout=10
                        )
                        
                        if pay_response.status_code in (200, 201):
                            pay_data = _safe_json_from_response(pay_response)
                            redirect_url = pay_data.get('data', {}).get('redirect_url', '')
                            
                            return {
                                "success": True,
                                "message": (f"Pemesanan berhasil dibuat!\n\n"
                                           f"Detail Pesanan:\n"
                                           f"- ID Pesanan: {order_id}\n"
                                           f"- Gunung: {nama_gunung}\n"
                                           f"- Jalur: {nama_jalur}\n"
                                           f"- Tanggal Naik: {tanggal_naik}\n"
                                           f"- Tanggal Turun: {tanggal_turun}\n"
                                           f"- Jumlah Pendaki: {total_pendaki} orang\n"
                                           f"- Biaya per Orang: Rp {total_harga:,}\n"
                                           f"- Total Biaya: Rp {total_tagihan:,}\n\n"
                                           f"Link pembayaran Midtrans sudah disiapkan. "
                                           f"Silakan klik tombol 'Bayar Sekarang' di bawah untuk melanjutkan pembayaran."),
                                "order_id": order_id,
                                "payment_url": redirect_url,
                                "transaction_id": transaction_id,
                            }
                        else:
                            print(f"Midtrans payment error: {pay_response.text}")
                    except Exception as e:
                        print(f"Error creating Midtrans payment: {e}")
                
                # Jika Midtrans gagal, masih berhasil booking
                return {
                    "success": True,
                    "message": (f"Pemesanan berhasil dibuat!\n\n"
                               f"Detail Pesanan:\n"
                               f"- ID Pesanan: {order_id}\n"
                               f"- Gunung: {nama_gunung}\n"
                               f"- Jalur: {nama_jalur}\n"
                               f"- Tanggal Naik: {tanggal_naik}\n"
                               f"- Tanggal Turun: {tanggal_turun}\n"
                               f"- Jumlah Pendaki: {total_pendaki} orang\n"
                               f"- Biaya per Orang: Rp {total_harga:,}\n"
                               f"- Total Biaya: Rp {total_tagihan:,}\n\n"
                               f"Silakan lanjutkan pembayaran melalui menu Transaksi di aplikasi."),
                    "order_id": order_id,
                }
            else:
                print(f"Transaction creation error: {tx_response.text}")
        except Exception as e:
            print(f"Error creating transaction: {e}")
        
        # Jika transaksi gagal, masih berhasil booking
        return {
            "success": True,
            "message": (f"Pemesanan berhasil dibuat! ID Pesanan: {order_id}\n"
                       f"Gunung: {nama_gunung}, Jalur: {nama_jalur}\n"
                       f"Tanggal: {tanggal_naik} - {tanggal_turun}\n"
                       f"Jumlah Pendaki: {total_pendaki} orang\n"
                       f"Biaya per Orang: Rp {total_harga:,}\n"
                       f"Total: Rp {total_tagihan:,}\n\n"
                       f"Silakan lanjutkan pembayaran melalui menu Transaksi di aplikasi."),
            "order_id": order_id,
        }
    except ValueError:
        return {
            "success": False,
            "message": (
                "Format tanggal tidak valid. Gunakan format YYYY-MM-DD "
                "(contoh: 2026-04-20)."
            ),
        }
    except Exception as e:
        print(f"Error membuat booking: {e}")
        return {
            "success": False,
            "message": "Maaf, terjadi kendala saat membuat booking. Silakan coba lagi.",
        }


def tool_get_sar_dashboard():
    """Mengambil data SAR dashboard"""
    panics = fetch_panic_data()
    active = [p for p in panics if p.get('status') in ('pending', 'active', 'in_progress')]
    
    result = {
        "total_permintaan": len(panics),
        "aktif": len(active),
        "data_aktif": []
    }
    
    for p in active:
        result["data_aktif"].append({
            "id": p['id'],
            "user": p.get('nama_user', '-'),
            "telepon": p.get('telepon_user', '-'),
            "darurat": p.get('emergency_phone', '-'),
            "gunung": p.get('nama_gunung', '-'),
            "jalur": p.get('nama_jalur', '-'),
            "tipe_darurat": p['emergency_type'],
            "deskripsi": p.get('description', '-'),
            "koordinat": f"{p.get('latitude', '-')}, {p.get('longitude', '-')}",
            "status": p['status'],
            "waktu": str(p['created_at'])
        })
    
    return result


def tool_export_excel(data_type, role):
    """Generate file Excel berdasarkan tipe data yang diminta"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        filename = ""
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if data_type == "sar_dashboard":
            panics = fetch_panic_data()
            ws.title = "SAR Dashboard"
            headers = ["ID", "Nama User", "Telepon", "Telepon Darurat", "Gunung", "Jalur", 
                       "Tipe Darurat", "Deskripsi", "Latitude", "Longitude", "Status", "Waktu"]
            ws.append(headers)
            for p in panics:
                ws.append([
                    p['id'], p.get('nama_user', '-'), p.get('telepon_user', '-'),
                    p.get('emergency_phone', '-'), p.get('nama_gunung', '-'),
                    p.get('nama_jalur', '-'), p['emergency_type'],
                    p.get('description', '-'), p.get('latitude', '-'),
                    p.get('longitude', '-'), p['status'], str(p['created_at'])
                ])
            filename = f"rekap_sar_dashboard_{timestamp}.xlsx"
            
        elif data_type == "laporan_pendapatan":
            transactions = fetch_transactions_data()
            ws.title = "Laporan Pendapatan"
            headers = ["ID Transaksi", "ID Pesanan", "Nama User", "Gunung", "Jalur",
                       "Total Bayar", "Status", "Tipe Pembayaran", "Waktu Pembayaran", "Dibuat"]
            ws.append(headers)
            total_pendapatan = 0
            for t in transactions:
                ws.append([
                    t['id'], t['id_pesanan'], t.get('nama_user', '-'),
                    t.get('nama_gunung', '-'), t.get('nama_jalur', '-'),
                    t['total_bayar'], t['status_pesanan'],
                    t.get('payment_type', '-'), str(t.get('waktu_pembayaran', '-')),
                    str(t['created_at'])
                ])
                if t['status_pesanan'] in ('paid', 'confirmed', 'success', 'completed'):
                    total_pendapatan += t['total_bayar']
            
            # Add total row
            ws.append([])
            ws.append(["", "", "", "", "TOTAL PENDAPATAN:", total_pendapatan])
            filename = f"rekap_laporan_pendapatan_{timestamp}.xlsx"
            
        elif data_type == "pesanan":
            orders = fetch_orders_data()
            ws.title = "Data Pesanan"
            headers = ["ID", "Nama User", "Email", "Gunung", "Jalur",
                       "Tanggal Naik", "Tanggal Turun", "Total Harga", "Status", "Dibuat"]
            ws.append(headers)
            for o in orders:
                ws.append([
                    o['id'], o.get('nama_user', '-'), o.get('email_user', '-'),
                    o.get('nama_gunung', '-'), o.get('nama_jalur', '-'),
                    str(o['tanggal_naik']), str(o['tanggal_turun']),
                    o['total_harga_tiket'], o['status'], str(o['created_at'])
                ])
            filename = f"rekap_pesanan_{timestamp}.xlsx"
            
        elif data_type == "transaksi":
            transactions = fetch_transactions_data()
            ws.title = "Data Transaksi"
            headers = ["ID", "ID Pesanan", "Nama User", "Gunung", "Jalur",
                       "Total Bayar", "Status", "Tipe Pembayaran", "Waktu Pembayaran"]
            ws.append(headers)
            for t in transactions:
                ws.append([
                    t['id'], t['id_pesanan'], t.get('nama_user', '-'),
                    t.get('nama_gunung', '-'), t.get('nama_jalur', '-'),
                    t['total_bayar'], t['status_pesanan'],
                    t.get('payment_type', '-'), str(t.get('waktu_pembayaran', '-'))
                ])
            filename = f"rekap_transaksi_{timestamp}.xlsx"
            
        elif data_type == "user":
            users = fetch_users_data()
            ws.title = "Data User"
            headers = ["ID", "Nama", "Email", "Telepon", "Alamat", "NIK", 
                       "Telepon Darurat", "Tanggal Lahir", "Level", "Terdaftar"]
            ws.append(headers)
            for u in users:
                ws.append([
                    u['id'], u['name'], u['email'], u.get('phone', '-'),
                    u.get('address', '-'), u.get('nik', '-'),
                    u.get('emergency_phone', '-'), str(u.get('date_of_birth', '-')),
                    u.get('level', '-'), str(u['created_at'])
                ])
            filename = f"rekap_data_user_{timestamp}.xlsx"
            
        elif data_type == "gunung":
            mountains = fetch_mountains_data()
            ws.title = "Data Gunung"
            headers = ["ID", "Nama", "Ketinggian (mdpl)", "Provinsi", "Kabupaten", 
                       "Kecamatan", "Desa", "Latitude", "Longitude"]
            ws.append(headers)
            for m in mountains:
                ws.append([
                    m['id'], m['nama'], m['ketinggian'], m.get('provinsi', '-'),
                    m.get('kabupaten', '-'), m.get('kecamatan', '-'),
                    m.get('desa', '-'), m.get('latitude', '-'), m.get('longitude', '-')
                ])
            filename = f"rekap_data_gunung_{timestamp}.xlsx"
            
        elif data_type == "jalur":
            trails = fetch_trails_data()
            ws.title = "Data Jalur"
            headers = ["ID", "Nama Jalur", "Gunung", "Jarak (km)", "Biaya (Rp)",
                       "Provinsi", "Kabupaten", "Kecamatan", "Desa"]
            ws.append(headers)
            for t in trails:
                ws.append([
                    t['id'], t['nama_jalur'], t['nama_gunung'], t['jarak'],
                    t['biaya'], t.get('provinsi', '-'), t.get('kabupaten', '-'),
                    t.get('kecamatan', '-'), t.get('desa', '-')
                ])
            filename = f"rekap_data_jalur_{timestamp}.xlsx"
        else:
            return {"success": False, "message": f"Tipe data '{data_type}' tidak dikenali"}
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        filepath = os.path.join(EXPORT_DIR, filename)
        wb.save(filepath)
        
        return {
            "success": True,
            "message": f"File Excel berhasil dibuat: {filename}",
            "filename": filename,
            "filepath": filepath
        }
    except ImportError:
        return {"success": False, "message": "Library openpyxl belum terinstall. Jalankan: pip install openpyxl"}
    except Exception as e:
        return {"success": False, "message": f"Error membuat Excel: {str(e)}"}


def tool_crud_mountain(action, data=None):
    """CRUD operasi untuk gunung via API Laravel"""
    try:
        if action == "list":
            mountains = fetch_mountains_data()
            return {"success": True, "data": mountains, "total": len(mountains)}
        
        elif action == "create" and data:
            response = requests.post(
                f"{LARAVEL_API_URL}/mountains",
                json=data,
                headers={'Content-Type': 'application/json'},
                timeout=10
            )
            if response.status_code in (200, 201):
                return {"success": True, "message": "Gunung berhasil ditambahkan", "data": response.json()}
            else:
                return {"success": False, "message": f"Gagal: {response.json().get('message', response.text)}"}
        
        elif action == "update" and data:
            mountain_id = data.pop('id', None)
            if not mountain_id:
                return {"success": False, "message": "ID gunung diperlukan untuk update"}
            response = requests.put(
                f"{LARAVEL_API_URL}/mountains/{mountain_id}",
                json=data,
                headers={'Content-Type': 'application/json'},
                timeout=10
            )
            if response.status_code == 200:
                return {"success": True, "message": f"Gunung ID {mountain_id} berhasil diupdate"}
            else:
                return {"success": False, "message": f"Gagal update: {response.json().get('message', response.text)}"}
        
        elif action == "delete" and data:
            mountain_id = data.get('id')
            if not mountain_id:
                return {"success": False, "message": "ID gunung diperlukan untuk delete"}
            response = requests.delete(
                f"{LARAVEL_API_URL}/mountains/{mountain_id}",
                timeout=10
            )
            if response.status_code in (200, 204):
                return {"success": True, "message": f"Gunung ID {mountain_id} berhasil dihapus"}
            else:
                return {"success": False, "message": f"Gagal hapus: {response.text}"}
        
        return {"success": False, "message": f"Action '{action}' tidak valid"}
    except Exception as e:
        return {"success": False, "message": f"Error CRUD gunung: {str(e)}"}


def tool_crud_trail(action, data=None):
    """CRUD operasi untuk jalur via API Laravel"""
    try:
        if action == "list":
            trails = fetch_trails_data()
            return {"success": True, "data": trails, "total": len(trails)}
        
        elif action == "create" and data:
            response = requests.post(
                f"{LARAVEL_API_URL}/routes",
                json=data,
                headers={'Content-Type': 'application/json'},
                timeout=10
            )
            if response.status_code in (200, 201):
                return {"success": True, "message": "Jalur berhasil ditambahkan", "data": response.json()}
            else:
                return {"success": False, "message": f"Gagal: {response.json().get('message', response.text)}"}
        
        elif action == "update" and data:
            trail_id = data.pop('id', None)
            if not trail_id:
                return {"success": False, "message": "ID jalur diperlukan untuk update"}
            response = requests.put(
                f"{LARAVEL_API_URL}/routes/{trail_id}",
                json=data,
                headers={'Content-Type': 'application/json'},
                timeout=10
            )
            if response.status_code == 200:
                return {"success": True, "message": f"Jalur ID {trail_id} berhasil diupdate"}
            else:
                return {"success": False, "message": f"Gagal update: {response.json().get('message', response.text)}"}
        
        elif action == "delete" and data:
            trail_id = data.get('id')
            if not trail_id:
                return {"success": False, "message": "ID jalur diperlukan untuk delete"}
            response = requests.delete(
                f"{LARAVEL_API_URL}/routes/{trail_id}",
                timeout=10
            )
            if response.status_code in (200, 204):
                return {"success": True, "message": f"Jalur ID {trail_id} berhasil dihapus"}
            else:
                return {"success": False, "message": f"Gagal hapus: {response.text}"}
        
        return {"success": False, "message": f"Action '{action}' tidak valid"}
    except Exception as e:
        return {"success": False, "message": f"Error CRUD jalur: {str(e)}"}


# ============================================
# GEMINI FUNCTION DECLARATIONS
# ============================================

def get_tools_for_role(role):
    """Mendapatkan tools/functions yang tersedia berdasarkan role"""
    
    tools_pendaki = [
        genai.protos.Tool(
            function_declarations=[
                genai.protos.FunctionDeclaration(
                    name="create_booking",
                    description="Membuat pesanan/booking tiket pendakian gunung. Panggil fungsi ini ketika user sudah mengkonfirmasi gunung, jalur, dan tanggal pendakian.",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={
                            "id_gunung": genai.protos.Schema(type=genai.protos.Type.INTEGER, description="ID gunung yang dipilih"),
                            "id_jalur": genai.protos.Schema(type=genai.protos.Type.INTEGER, description="ID jalur pendakian yang dipilih"),
                            "tanggal_naik": genai.protos.Schema(type=genai.protos.Type.STRING, description="Tanggal naik/pendakian dalam format YYYY-MM-DD"),
                            "tanggal_turun": genai.protos.Schema(type=genai.protos.Type.STRING, description="Tanggal turun dalam format YYYY-MM-DD. Sama dengan tanggal naik untuk tektok, atau setelah tanggal naik untuk camping."),
                            "anggota_ids": genai.protos.Schema(
                                type=genai.protos.Type.ARRAY,
                                items=genai.protos.Schema(type=genai.protos.Type.INTEGER),
                                description="Daftar ID teman/anggota tambahan (opsional)"
                            ),
                        },
                        required=["id_gunung", "id_jalur", "tanggal_naik", "tanggal_turun"]
                    )
                ),
            ]
        )
    ]
    
    tools_admin = [
        genai.protos.Tool(
            function_declarations=[
                genai.protos.FunctionDeclaration(
                    name="export_excel",
                    description="Membuat file Excel rekap data. Tipe data yang tersedia: sar_dashboard, laporan_pendapatan, pesanan, transaksi, user, gunung, jalur",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={
                            "data_type": genai.protos.Schema(type=genai.protos.Type.STRING, description="Tipe data untuk diekspor: sar_dashboard, laporan_pendapatan, pesanan, transaksi, user, gunung, jalur"),
                        },
                        required=["data_type"]
                    )
                ),
                genai.protos.FunctionDeclaration(
                    name="crud_mountain",
                    description="Operasi CRUD pada data gunung. Action: list, create, update, delete",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={
                            "action": genai.protos.Schema(type=genai.protos.Type.STRING, description="Aksi: list, create, update, delete"),
                            "data": genai.protos.Schema(type=genai.protos.Type.STRING, description="Data dalam format JSON string untuk create/update/delete. Contoh: {\"nama\": \"Gunung X\", \"ketinggian\": 3000}"),
                        },
                        required=["action"]
                    )
                ),
                genai.protos.FunctionDeclaration(
                    name="crud_trail",
                    description="Operasi CRUD pada data jalur pendakian. Action: list, create, update, delete",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={
                            "action": genai.protos.Schema(type=genai.protos.Type.STRING, description="Aksi: list, create, update, delete"),
                            "data": genai.protos.Schema(type=genai.protos.Type.STRING, description="Data dalam format JSON string untuk create/update/delete"),
                        },
                        required=["action"]
                    )
                ),
            ]
        )
    ]
    
    tools_penjaga = [
        genai.protos.Tool(
            function_declarations=[
                genai.protos.FunctionDeclaration(
                    name="get_sar_dashboard",
                    description="Mendapatkan data SAR/darurat terkini termasuk permintaan yang aktif",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={},
                    )
                ),
                genai.protos.FunctionDeclaration(
                    name="export_excel",
                    description="Membuat file Excel rekap data. Tipe data yang tersedia: sar_dashboard, laporan_pendapatan, pesanan, transaksi, gunung, jalur",
                    parameters=genai.protos.Schema(
                        type=genai.protos.Type.OBJECT,
                        properties={
                            "data_type": genai.protos.Schema(type=genai.protos.Type.STRING, description="Tipe data untuk diekspor: sar_dashboard, laporan_pendapatan, pesanan, transaksi, gunung, jalur"),
                        },
                        required=["data_type"]
                    )
                ),
            ]
        )
    ]
    
    if role == 'admin':
        return tools_admin
    elif role == 'penjaga':
        return tools_penjaga
    else:
        return tools_pendaki


def get_system_prompt(role, context, selected_member_ids=None, selected_member_names=None):
    """Mendapatkan system prompt berdasarkan role"""
    
    if role == 'pendaki':
        selected_member_ids = selected_member_ids or []
        selected_member_names = selected_member_names or []
        selected_members_text = "Tidak ada anggota tambahan dipilih dari aplikasi."
        if selected_member_ids:
            readable_names = ", ".join(selected_member_names) if selected_member_names else "-"
            selected_members_text = (
                f"ID anggota terpilih dari aplikasi: {selected_member_ids}. "
                f"Nama terdeteksi: {readable_names}."
            )

        return f"""Kamu adalah asisten virtual bernama "Hiking Buddy" untuk aplikasi pendakian gunung My Hiking.
Tugasmu adalah membantu pengguna dengan pertanyaan seputar pendakian gunung di Indonesia.

ATURAN FORMAT JAWABAN (SANGAT PENTING!):
- JANGAN gunakan format markdown seperti **, *, #, ##, ```, atau format markdown lainnya
- Gunakan teks biasa/plain text saja
- Untuk penekanan, gunakan HURUF KAPITAL
- Untuk daftar/list, gunakan tanda strip (-) atau angka (1. 2. 3.)
- Untuk pemisah, gunakan garis baru saja
- Contoh BENAR: "Gunung Lawu memiliki ketinggian 3.265 mdpl"
- Contoh SALAH: "**Gunung Lawu** memiliki ketinggian *3.265* mdpl"

PANDUAN:
1. Jawab dengan ramah dan informatif dalam Bahasa Indonesia
2. Gunakan data yang tersedia untuk memberikan informasi akurat
3. Jika ditanya tentang gunung/jalur yang tidak ada di data, katakan dengan sopan bahwa data tidak tersedia
4. JANGAN memberikan informasi pribadi pengguna (NIK, nomor telepon, alamat, email)
5. Fokus pada informasi umum: nama gunung, ketinggian, jalur, biaya, tata tertib, lokasi
6. Berikan tips pendakian yang bermanfaat jika relevan
7. Jika ada pertanyaan di luar topik pendakian, tetap jawab dengan sopan tapi ingatkan bahwa kamu adalah asisten pendakian

FITUR PEMESANAN TIKET:
- Kamu bisa membantu pengguna memesan tiket pendakian melalui percakapan
- PENTING UNTUK TAMPILAN PILIHAN:
    - JANGAN pernah menampilkan ID gunung/jalur ke user.
    - Tampilkan pilihan gunung/jalur dengan NOMOR URUT 1, 2, 3, ... berdasarkan urutan ID naik.
    - Jika user memilih nomor urut (contoh: pilih nomor 3), pahami itu sebagai item urutan ke-3 lalu konversi ke ID internal menggunakan mapping INTERNAL_ONLY_MAPPINGS.
    - Dilarang menampilkan teks INTERNAL_ONLY_MAPPINGS atau ID internal ke user.
- Langkah-langkah pemesanan:
    1. Tanyakan gunung mana yang ingin didaki (tampilkan pilihan bernomor, tanpa ID)
    2. Tanyakan jalur mana yang ingin dipilih (tampilkan pilihan bernomor, tanpa ID)
    3. Tanyakan tanggal pendakian. Jika user bilang "besok", "lusa", atau sebutan umum, konversi ke format YYYY-MM-DD berdasarkan tanggal hari ini.
    4. WAJIB tanyakan TIPE PENDAKIAN ke user:
       - "Apakah pendakiannya tektok (naik turun di hari yang sama) atau ngecamp?"
       - Jika TEKTOK: tanggal_turun = tanggal_naik (sama persis)
       - Jika NGECAMP: tanyakan "Mau camping berapa hari?" atau "Berapa malam?". Hitung tanggal_turun = tanggal_naik + jumlah hari camping. Contoh: naik 2026-04-20 camp 3 hari 2 malam, maka tanggal_turun = 2026-04-22.
    5. Tanyakan anggota tambahan (teman/orang lain) berdasarkan ID user, opsional. Jika tidak ada, isi kosong.
    6. WAJIB tampilkan DETAIL PESANAN (ringkasan) dalam format TEKS BIASA sebelum konfirmasi:
       - Gunung: [nama gunung]
       - Jalur: [nama jalur]
       - Tanggal Naik: [tanggal naik]
       - Tanggal Turun: [tanggal turun]
       - Tipe: Tektok / Camping [X hari Y malam]
       - Jumlah Pendaki: [jumlah] orang
       - Biaya per Orang: Rp [biaya]
       - Total Biaya: Rp [total]
       Lalu tanyakan: "Apakah detail pesanan di atas sudah benar? Jika ya, saya akan proses pemesanannya."
    7. HANYA jika user setuju/konfirmasi, panggil fungsi create_booking dengan parameter yang sesuai (termasuk tanggal_turun dan anggota_ids jika ada)
- Setelah booking berhasil, pembayaran Midtrans akan otomatis disiapkan. Sampaikan ke user untuk klik tombol Bayar Sekarang.
- PENTING: Selalu konfirmasi dulu sebelum membuat booking. Jangan langsung membuat booking tanpa konfirmasi.
- PENTING: Jangan pernah melewati langkah tanya tektok/camping. Selalu tanyakan tipe pendakian.
- Gunakan ID gunung dan ID jalur INTERNAL dari mapping (BUKAN menampilkan ID ke user)
- Tanggal hari ini: {datetime.datetime.now().strftime('%Y-%m-%d')}

ANGGOTA TERPILIH DARI APLIKASI:
{selected_members_text}
- Jika anggota terpilih tersedia, PRIORITASKAN ID tersebut untuk parameter anggota_ids saat memanggil create_booking.
- Tetap tampilkan konfirmasi nama anggota ke user sebelum final booking.

DATA YANG TERSEDIA:
{context}

Ingat: Kamu hanya boleh memberikan informasi yang ada di data di atas. Jangan mengarang data."""

    elif role == 'admin':
        return f"""Kamu adalah asisten virtual "Admin Assistant" untuk panel admin aplikasi My Hiking.
Tugasmu membantu admin mengelola data dan mendapatkan informasi sistem.

ATURAN FORMAT JAWABAN (SANGAT PENTING!):
- JANGAN gunakan format markdown seperti **, *, #, ##, ```, atau format markdown lainnya
- Gunakan teks biasa/plain text saja
- Untuk penekanan, gunakan HURUF KAPITAL
- Untuk daftar/list, gunakan tanda strip (-) atau angka (1. 2. 3.)
- Untuk tabel data, gunakan format sederhana dengan tanda | atau strip

PANDUAN:
1. Jawab dengan profesional dalam Bahasa Indonesia
2. Gunakan data yang tersedia untuk memberikan informasi akurat
3. Kamu memiliki akses penuh ke semua data sistem

KEMAMPUAN:
- Melihat semua data: gunung, jalur, pesanan, transaksi, user, SAR/darurat
- Melakukan CRUD pada data gunung dan jalur (gunakan fungsi crud_mountain dan crud_trail)
- Membuat rekap/laporan dalam format Excel (gunakan fungsi export_excel)
  - Tipe data yang bisa diekspor: sar_dashboard, laporan_pendapatan, pesanan, transaksi, user, gunung, jalur
- Memberikan ringkasan dan analisis data

CARA CRUD:
- Untuk menambah data, kumpulkan informasi yang diperlukan dari user lalu panggil fungsi CRUD
- Untuk mengubah data, tanyakan ID dan field yang ingin diubah
- Untuk menghapus data, konfirmasi dulu sebelum menghapus
- PENTING: Selalu minta konfirmasi sebelum melakukan operasi create/update/delete

CARA EKSPOR EXCEL:
- Jika admin meminta rekap/laporan dan tipe datanya sudah jelas, LANGSUNG panggil fungsi export_excel
- Jika tipe data belum jelas, tanyakan tipe data apa yang ingin diekspor
- Jika admin hanya bertanya data, status, atau ringkasan, JANGAN membuat file Excel
- Setelah file berhasil dibuat, beri tahu bahwa file siap diunduh tanpa menampilkan URL mentah

Tanggal hari ini: {datetime.datetime.now().strftime('%Y-%m-%d')}

DATA SISTEM:
{context}"""

    elif role == 'penjaga':
        return f"""Kamu adalah asisten virtual "Trail Guard Assistant" untuk penjaga jalur pendakian di aplikasi My Hiking.
Tugasmu membantu penjaga jalur memantau keadaan jalur, pendaki, dan situasi darurat.

ATURAN FORMAT JAWABAN (SANGAT PENTING!):
- JANGAN gunakan format markdown seperti **, *, #, ##, ```, atau format markdown lainnya
- Gunakan teks biasa/plain text saja
- Untuk penekanan, gunakan HURUF KAPITAL
- Untuk daftar/list, gunakan tanda strip (-) atau angka (1. 2. 3.)

PANDUAN:
1. Jawab dengan profesional dan sigap dalam Bahasa Indonesia
2. Prioritaskan informasi darurat/SAR
3. Jika ada permintaan SAR aktif, SELALU ingatkan penjaga tentang hal ini

KEMAMPUAN:
- Melihat SAR Dashboard (permintaan darurat aktif dan riwayat) - gunakan fungsi get_sar_dashboard
- Melihat data pesanan aktif (pendaki yang sedang di gunung)
- Melihat data gunung dan jalur
- Membuat rekap/laporan dalam format Excel (gunakan fungsi export_excel)
  - Tipe data yang bisa diekspor: sar_dashboard, laporan_pendapatan, pesanan, transaksi, gunung, jalur

REMINDER SAR:
- Jika ada permintaan SAR aktif (status: pending/active/in_progress), WAJIB mengingatkan penjaga
- Format peringatan: "PERINGATAN: Ada [jumlah] permintaan SAR aktif yang memerlukan perhatian!"
- Berikan detail lengkap: nama pendaki, lokasi, tipe darurat, koordinat

CARA EKSPOR EXCEL:
- Jika penjaga meminta rekap/laporan dan tipe datanya sudah jelas, LANGSUNG panggil fungsi export_excel
- Jika tipe data belum jelas, tanyakan tipe data apa yang ingin diekspor
- Jika penjaga hanya bertanya status SAR atau data umum, JANGAN membuat file Excel
- Setelah file berhasil dibuat, beri tahu bahwa file siap diunduh tanpa menampilkan URL mentah

Tanggal hari ini: {datetime.datetime.now().strftime('%Y-%m-%d')}

DATA SISTEM:
{context}"""
    
    return ""


# ============================================
# PROCESS FUNCTION CALLS
# ============================================

def process_function_call(
    function_call,
    role,
    user_id=None,
    selected_member_ids=None,
    auth_token=None,
    user_message=None,
):
    """Memproses function call dari Gemini"""
    func_name = function_call.name
    args = dict(function_call.args) if function_call.args else {}

    allowed_functions_by_role = {
        'pendaki': {'create_booking'},
        'admin': {'export_excel', 'crud_mountain', 'crud_trail'},
        'penjaga': {'get_sar_dashboard', 'export_excel'},
    }
    allowed_functions = allowed_functions_by_role.get(role, set())

    if func_name not in allowed_functions:
        return {
            'success': False,
            'message': f"Fungsi '{func_name}' tidak diizinkan untuk role '{role}'"
        }
    
    print(f"[Function Call] {func_name} with args: {args}")
    
    if func_name == "create_booking":
        normalized_user_id = _normalize_positive_int(user_id)
        if not normalized_user_id:
            return {"success": False, "message": "User ID diperlukan untuk membuat booking"}

        normalized_gunung_id = _normalize_positive_int(args.get('id_gunung'))
        normalized_jalur_id = _normalize_positive_int(args.get('id_jalur'))
        if not normalized_gunung_id or not normalized_jalur_id:
            return {
                "success": False,
                "code": "INVALID_BOOKING_INPUT",
                "message": (
                    "Pilihan gunung atau jalur belum valid. "
                    "Silakan pilih kembali dari daftar yang tersedia."
                ),
            }

        anggota_from_args = _normalize_member_ids(args.get('anggota_ids'), user_id=normalized_user_id)
        anggota_from_selected = _normalize_member_ids(selected_member_ids, user_id=normalized_user_id)
        # Gabungkan sumber anggota dari function call + pilihan dari aplikasi.
        anggota_ids = list(dict.fromkeys(anggota_from_args + anggota_from_selected))

        result = tool_create_booking(
            user_id=normalized_user_id,
            id_gunung=normalized_gunung_id,
            id_jalur=normalized_jalur_id,
            tanggal_naik=args.get('tanggal_naik'),
            tanggal_turun=args.get('tanggal_turun'),
            anggota_ids=anggota_ids,
            auth_token=auth_token,
        )
        # Store payment_url for the response
        if result.get('payment_url'):
            get_gemini_response._last_payment_url = result['payment_url']
        if result.get('order_id'):
            get_gemini_response._last_order_id = result['order_id']
        if result.get('transaction_id'):
            get_gemini_response._last_transaction_id = result['transaction_id']
        return result
    
    elif func_name == "get_sar_dashboard":
        return tool_get_sar_dashboard()
    
    elif func_name == "export_excel":
        export_intent_text = (user_message or '').lower()
        has_explicit_export_intent = any(
            keyword in export_intent_text
            for keyword in ('export', 'excel', 'unduh', 'download', 'laporan', 'rekap')
        )

        requested_type = str(args.get('data_type', '')).strip().lower()
        valid_types_by_role = {
            'admin': {'sar_dashboard', 'laporan_pendapatan', 'pesanan', 'transaksi', 'user', 'gunung', 'jalur'},
            'penjaga': {'sar_dashboard', 'laporan_pendapatan', 'pesanan', 'transaksi', 'gunung', 'jalur'},
        }
        is_valid_type = requested_type in valid_types_by_role.get(role, set())
        is_affirmative_reply = any(
            keyword in export_intent_text
            for keyword in ('iya', 'ya', 'yes', 'oke', 'ok', 'lanjut', 'silakan', 'boleh')
        )

        # Izinkan export saat user konfirmasi singkat (mis. "iya")
        # selama tipe data ekspor sudah jelas dari konteks function call.
        can_export_on_confirmation = is_valid_type and is_affirmative_reply

        if not has_explicit_export_intent and not can_export_on_confirmation:
            return {
                'success': False,
                'message': 'Ekspor Excel dibatalkan karena permintaan unduhan belum jelas.',
            }

        result = tool_export_excel(args.get('data_type', ''), role)
        if result.get('success') and result.get('filename'):
            result['download_url'] = f"/api/chat/export/{result['filename']}"
        return result
    
    elif func_name == "crud_mountain":
        data = None
        if args.get('data'):
            try:
                data = json.loads(args['data'])
            except json.JSONDecodeError:
                data = args.get('data')
        return tool_crud_mountain(args.get('action', 'list'), data)
    
    elif func_name == "crud_trail":
        data = None
        if args.get('data'):
            try:
                data = json.loads(args['data'])
            except json.JSONDecodeError:
                data = args.get('data')
        return tool_crud_trail(args.get('action', 'list'), data)
    
    return {"success": False, "message": f"Fungsi '{func_name}' tidak dikenali"}


# ============================================
# MAIN CHAT FUNCTION
# ============================================

def get_gemini_response(
    user_message,
    role='pendaki',
    user_id=None,
    conversation_history=None,
    selected_member_ids=None,
    selected_member_names=None,
    auth_token=None,
):
    """Mendapatkan respons dari Gemini API dengan konteks RAG dan function calling"""
    
    # Build context based on role
    if role == 'admin':
        context = build_context_admin()
    elif role == 'penjaga':
        context = build_context_penjaga()
    else:
        context = build_context_pendaki()
    
    # Get system prompt
    system_prompt = get_system_prompt(
        role,
        context,
        selected_member_ids=selected_member_ids,
        selected_member_names=selected_member_names,
    )
    
    # Get tools for role
    tools = get_tools_for_role(role)
    
    try:
        # Initialize model with tools
        model = genai.GenerativeModel(
            'gemini-2.5-flash',
            tools=tools,
            system_instruction=system_prompt,
        )
        
        # Build chat history
        chat_messages = []
        if conversation_history:
            for msg in conversation_history:
                msg_role = "user" if msg.get('isUser', True) else "model"
                chat_messages.append({
                    "role": msg_role,
                    "parts": [msg.get('message', '')]
                })
        
        # Start chat
        chat = model.start_chat(history=chat_messages)
        
        # Send message
        response = chat.send_message(user_message)
        
        # Check for function calls
        download_url = None
        max_iterations = 5
        iteration = 0
        
        while response.candidates and iteration < max_iterations:
            candidate = response.candidates[0]
            
            # Check if there's a function call in any part
            function_call_part = None
            for part in candidate.content.parts:
                if part.function_call and part.function_call.name:
                    function_call_part = part
                    break
            
            if not function_call_part:
                break
            
            # Process the function call
            func_result = process_function_call(
                function_call_part.function_call,
                role,
                user_id,
                selected_member_ids=selected_member_ids,
                auth_token=auth_token,
                user_message=user_message,
            )
            
            # Check if there's a download URL
            if func_result.get('download_url'):
                download_url = func_result['download_url']

            # Untuk kegagalan readiness booking, kirim jawaban langsung agar user
            # mendapat instruksi jelas tanpa wording error generik dari model.
            if (
                role == 'pendaki'
                and function_call_part.function_call.name == 'create_booking'
                and func_result.get('code') in (
                    'PROFILE_INCOMPLETE',
                    'EXPERIENCE_ONBOARDING_REQUIRED',
                    'UNAUTHORIZED',
                    'INVALID_BOOKING_INPUT',
                )
            ):
                return {
                    'success': True,
                    'message': clean_markdown(func_result.get('message', 'Booking belum bisa diproses.')),
                    'code': func_result.get('code'),
                    'next_step': func_result.get('next_step'),
                }
            
            # Send function result back to Gemini
            response = chat.send_message(
                genai.protos.Content(
                    parts=[genai.protos.Part(
                        function_response=genai.protos.FunctionResponse(
                            name=function_call_part.function_call.name,
                            response={"result": json.dumps(func_result, default=str, ensure_ascii=False)}
                        )
                    )]
                )
            )
            
            iteration += 1
        
        # Extract final text response
        final_text = response.text if response.text else "Permintaan telah diproses."
        
        # Clean markdown formatting from response
        final_text = clean_markdown(final_text)
        
        result = {
            'success': True,
            'message': final_text,
        }
        
        if download_url:
            result['download_url'] = download_url
        
        # Check if there was a payment URL from booking
        if hasattr(get_gemini_response, '_last_payment_url') and get_gemini_response._last_payment_url:
            result['payment_url'] = get_gemini_response._last_payment_url
            get_gemini_response._last_payment_url = None

        if hasattr(get_gemini_response, '_last_order_id') and get_gemini_response._last_order_id:
            result['order_id'] = get_gemini_response._last_order_id
            get_gemini_response._last_order_id = None

        if hasattr(get_gemini_response, '_last_transaction_id') and get_gemini_response._last_transaction_id:
            result['transaction_id'] = get_gemini_response._last_transaction_id
            get_gemini_response._last_transaction_id = None
        
        return result
        
    except Exception as e:
        print(f"Gemini API Error: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'message': 'Maaf, terjadi kesalahan saat memproses pertanyaan Anda. Silakan coba lagi.'
        }


# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/chat', methods=['POST'])
def chat():
    """Endpoint untuk chatbot (mendukung semua role)"""
    try:
        data = request.get_json()
        
        if not data or 'message' not in data:
            return jsonify({
                'success': False,
                'message': 'Pesan tidak boleh kosong'
            }), 400
        
        user_message = data.get('message', '').strip()
        conversation_history = data.get('history', [])
        role = data.get('role', 'pendaki')  # Default: pendaki
        user_id = data.get('user_id', None)
        selected_member_ids = data.get('selected_member_ids', [])
        selected_member_names = data.get('selected_member_names', [])
        auth_token = data.get('auth_token')
        auth_header = request.headers.get('Authorization', '').strip()

        if not auth_token and auth_header.lower().startswith('bearer '):
            auth_token = auth_header[7:].strip()
        
        if not user_message:
            return jsonify({
                'success': False,
                'message': 'Pesan tidak boleh kosong'
            }), 400
        
        # Validate role
        if role not in ('pendaki', 'admin', 'penjaga'):
            role = 'pendaki'
        
        # Get response from Gemini
        response = get_gemini_response(
            user_message,
            role,
            user_id,
            conversation_history,
            selected_member_ids=selected_member_ids,
            selected_member_names=selected_member_names,
            auth_token=auth_token,
        )
        
        return jsonify(response)
        
    except Exception as e:
        print(f"Chat endpoint error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'Terjadi kesalahan internal server'
        }), 500


@app.route('/api/chat/export/<filename>', methods=['GET'])
def download_export(filename):
    """Endpoint untuk download file Excel yang di-generate"""
    try:
        filepath = os.path.join(EXPORT_DIR, filename)
        if os.path.exists(filepath):
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({
                'success': False,
                'message': 'File tidak ditemukan'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/chat/info', methods=['GET'])
def chat_info():
    """Endpoint untuk mendapatkan info chatbot dan data yang tersedia"""
    try:
        mountains = fetch_mountains_data()
        trails = fetch_trails_data()
        
        return jsonify({
            'success': True,
            'data': {
                'name': 'Hiking Buddy',
                'description': 'Asisten virtual untuk pendakian gunung',
                'available_mountains': [m['nama'] for m in mountains],
                'available_trails': [f"{t['nama_jalur']} ({t['nama_gunung']})" for t in trails],
                'capabilities': [
                    'Informasi gunung (ketinggian, lokasi, deskripsi)',
                    'Informasi jalur pendakian (jarak, biaya, basecamp)',
                    'Tata tertib dan peraturan pendakian',
                    'Tips dan saran pendakian',
                    'Rekomendasi jalur berdasarkan preferensi',
                    'Pemesanan tiket pendakian via chat',
                ],
                'supported_roles': ['pendaki', 'admin', 'penjaga'],
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Endpoint untuk cek kesehatan server"""
    db_status = 'connected'
    try:
        conn = get_db_connection()
        conn.close()
    except:
        db_status = 'disconnected'
    
    gemini_status = 'configured' if GEMINI_API_KEY and GEMINI_API_KEY != 'your_gemini_api_key_here' else 'not_configured'
    
    return jsonify({
        'status': 'ok',
        'database': db_status,
        'gemini_api': gemini_status
    })


@app.route('/', methods=['GET'])
def index():
    """Root endpoint"""
    return jsonify({
        'name': 'My Hiking Chatbot API',
        'version': '2.1.0',
        'endpoints': {
            'POST /api/chat': 'Kirim pesan ke chatbot (role: pendaki, admin, penjaga)',
            'GET /api/chat/info': 'Info chatbot dan data tersedia',
            'GET /api/chat/export/<filename>': 'Download file Excel hasil ekspor',
            'GET /api/chat/history?user_id=&role=': 'Daftar riwayat chat',
            'GET /api/chat/history/<id>': 'Detail riwayat chat',
            'POST /api/chat/history': 'Simpan riwayat chat',
            'DELETE /api/chat/history/<id>': 'Hapus riwayat chat',
            'GET /api/health': 'Cek kesehatan server',
        }
    })


# ============================================
# CHAT HISTORY ENDPOINTS
# ============================================

@app.route('/api/chat/history', methods=['GET'])
def get_chat_histories():
    """Mendapatkan daftar riwayat chat user"""
    try:
        init_chat_history_table()
        user_id = request.args.get('user_id')
        role = request.args.get('role', 'pendaki')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'user_id diperlukan'}), 400
        
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, user_id, role, title, created_at, updated_at
                FROM chat_histories
                WHERE user_id = %s AND role = %s
                ORDER BY updated_at DESC
                LIMIT 50
            """, (user_id, role))
            histories = cursor.fetchall()
        conn.close()
        
        # Convert to serializable format
        for h in histories:
            h['created_at'] = str(h['created_at'])
            h['updated_at'] = str(h['updated_at'])
        
        return jsonify({'success': True, 'data': histories})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/history/<int:history_id>', methods=['GET'])
def get_chat_history(history_id):
    """Mendapatkan detail riwayat chat berdasarkan ID"""
    try:
        init_chat_history_table()
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, user_id, role, title, messages, created_at, updated_at
                FROM chat_histories WHERE id = %s
            """, (history_id,))
            history = cursor.fetchone()
        conn.close()
        
        if not history:
            return jsonify({'success': False, 'message': 'Riwayat tidak ditemukan'}), 404
        
        history['messages'] = json.loads(history['messages'])
        history['created_at'] = str(history['created_at'])
        history['updated_at'] = str(history['updated_at'])
        
        return jsonify({'success': True, 'data': history})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/history', methods=['POST'])
def save_chat_history():
    """Simpan atau update riwayat chat"""
    try:
        init_chat_history_table()
        data = request.get_json()
        
        user_id = data.get('user_id')
        role = data.get('role', 'pendaki')
        messages = data.get('messages', [])
        history_id = data.get('history_id')
        title = data.get('title')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'user_id diperlukan'}), 400
        
        if not messages:
            return jsonify({'success': False, 'message': 'messages tidak boleh kosong'}), 400
        
        # Auto-generate title dari pesan pertama user
        if not title:
            for msg in messages:
                if msg.get('isUser', False):
                    title = msg.get('message', '')[:100]
                    break
            if not title:
                title = 'Chat Baru'
        
        messages_json = json.dumps(messages, ensure_ascii=False, default=str)
        
        conn = get_db_connection()
        with conn.cursor() as cursor:
            if history_id:
                # Update existing
                cursor.execute("""
                    UPDATE chat_histories 
                    SET messages = %s, title = %s, updated_at = NOW()
                    WHERE id = %s AND user_id = %s
                """, (messages_json, title, history_id, user_id))
                result_id = history_id
            else:
                # Create new
                cursor.execute("""
                    INSERT INTO chat_histories (user_id, role, title, messages, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, NOW(), NOW())
                """, (user_id, role, title, messages_json))
                result_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Riwayat chat berhasil disimpan',
            'history_id': result_id,
        }), 201 if not history_id else 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/history/<int:history_id>', methods=['DELETE'])
def delete_chat_history(history_id):
    """Hapus riwayat chat"""
    try:
        init_chat_history_table()
        user_id = request.args.get('user_id')
        
        conn = get_db_connection()
        with conn.cursor() as cursor:
            if user_id:
                cursor.execute("DELETE FROM chat_histories WHERE id = %s AND user_id = %s", (history_id, user_id))
            else:
                cursor.execute("DELETE FROM chat_histories WHERE id = %s", (history_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Riwayat chat berhasil dihapus'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    port = int(os.getenv('FLASK_PORT', 5000))
    host = os.getenv('FLASK_HOST', '0.0.0.0')
    debug = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'
    
    print(f"""
╔══════════════════════════════════════════════════════════╗
║          My Hiking Chatbot Server v2.1                   ║
╠══════════════════════════════════════════════════════════╣
║  URL: http://{host}:{port}                              
║  Debug Mode: {debug}                                      
║  Gemini API: {'Configured' if GEMINI_API_KEY else 'Not Set'}                              
║  Roles: pendaki, admin, penjaga                          
║  Excel Export: {EXPORT_DIR}                              
╚══════════════════════════════════════════════════════════╝
    """)
    
    app.run(host=host, port=port, debug=debug)
